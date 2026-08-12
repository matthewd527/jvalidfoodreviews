#!/usr/bin/env python3
"""
Convert the hand-researched ratings spreadsheet into data/ratings.js.

This is deliberately NOT part of the daily job. The daily scrape can read view
counts off TikTok, but nobody can automate "what score did he say out loud" -
that comes from watching the videos. So the ranking is a curated dataset that
you re-import whenever the spreadsheet grows.

The two data files stay separate on purpose: scripts/update.py owns
data/site.js and rewrites it every day, and would happily clobber anything
hand-made living inside it.

Usage:
  python3 scripts/import_ratings.py ~/Downloads/jvalid_food_reviews.xlsx
"""

from __future__ import annotations

import json
import pathlib
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is needed for this one-off import:  pip3 install openpyxl")

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "ratings.js"

# Scores he actually said. Anything outside 1-5 is a joke score and is flagged
# rather than quietly clamped, because the jokes are the point of the account.
SCALE_MIN, SCALE_MAX = 1.0, 5.0


def tier(score: float | None) -> str:
    if score is None:
        return "unrated"
    if score > SCALE_MAX or score < SCALE_MIN:
        return "offscale"
    if score >= 4.7:
        return "elite"
    if score >= 4.0:
        return "great"
    if score >= 3.0:
        return "fine"
    return "rough"


def short_place(location: str | None) -> str:
    """'395 S Main St, New City, NY 10956' -> 'New City, NY'."""
    if not location:
        return ""
    loc = location.strip()
    if loc.lower().startswith("not stated"):
        return ""
    loc = re.sub(r"\s*\(.*?\)\s*", " ", loc).strip()          # drop parentheticals
    parts = [p.strip() for p in loc.split(",") if p.strip()]
    if not parts:
        return ""
    # Find the "NY 10956" / "NJ" chunk and pair it with whatever precedes it.
    for i, p in enumerate(parts):
        m = re.match(r"^([A-Z]{2})\b", p)
        if m and i > 0:
            return f"{parts[i - 1]}, {m.group(1)}"
    return ", ".join(parts[-2:]) if len(parts) > 1 else parts[0]


def main() -> int:
    src = pathlib.Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else None
    if not src or not src.exists():
        sys.exit("usage: python3 scripts/import_ratings.py <path to .xlsx>")

    wb = openpyxl.load_workbook(src, data_only=True)

    # ── per-item scores, grouped by video ────────────────────────────────────
    items: dict[int, list[dict]] = {}
    for row in wb["Food Ratings"].iter_rows(min_row=4, values_only=True):
        vid, _rest, item, rating, evidence = (list(row) + [None] * 5)[:5]
        if vid is None or not isinstance(rating, (int, float)):
            continue
        items.setdefault(int(vid), []).append({
            "item": str(item).strip() if item else "Unnamed item",
            "score": round(float(rating), 2),
            "note": str(evidence).strip() if evidence else "",
        })

    # ── caveats from the methodology sheet ───────────────────────────────────
    caveats: dict[int, list[str]] = {}
    seen_header = False
    for row in wb["Method & Gaps"].iter_rows(values_only=True):
        cells = list(row) + [None] * 4
        if cells[0] == "Video #":
            seen_header = True
            continue
        if seen_header and isinstance(cells[0], (int, float)):
            caveats.setdefault(int(cells[0]), []).append(
                f"{cells[2]}: {cells[3]}" if cells[2] else str(cells[3])
            )

    # ── one entry per video ──────────────────────────────────────────────────
    entries = []
    for row in wb["Video Summary"].iter_rows(min_row=7, values_only=True):
        n, posted, rest, loc, _foods, avg, overall, verif, tt, venue = (list(row) + [None] * 10)[:10]
        if n is None or not rest:
            continue
        n = int(n)
        score = round(float(avg), 3) if isinstance(avg, (int, float)) else None
        m = re.search(r"/video/(\d+)", str(tt or ""))

        entries.append({
            "n": n,
            "name": str(rest).strip(),
            "place": short_place(str(loc) if loc else ""),
            "address": (str(loc).strip() if loc and not str(loc).lower().startswith("not stated") else ""),
            "score": score,
            "tier": tier(score),
            "overall": round(float(overall), 2) if isinstance(overall, (int, float)) else None,
            "items": sorted(items.get(n, []), key=lambda i: -i["score"]),
            "posted": posted.strftime("%Y-%m-%d") if hasattr(posted, "strftime") else None,
            "videoId": m.group(1) if m else None,
            "caveats": caveats.get(n, []),
            "verified": str(verif).strip() if verif else "",
        })

    rated = [e for e in entries if e["score"] is not None]
    unrated = [e for e in entries if e["score"] is None]
    rated.sort(key=lambda e: (-e["score"], e["name"].lower()))

    # Competition ranking: equal scores share a rank, and the next rank skips.
    last, last_rank = None, 0
    for i, e in enumerate(rated, 1):
        if e["score"] != last:
            last, last_rank = e["score"], i
        e["rank"] = last_rank

    scores = [e["score"] for e in rated]
    payload = {
        "source": src.name,
        "totalVideos": len(entries),
        "ratedVideos": len(rated),
        "ratingEvents": sum(len(v) for v in items.values()),
        "best": rated[0]["score"] if rated else None,
        "worst": rated[-1]["score"] if rated else None,
        "average": round(sum(scores) / len(scores), 2) if scores else None,
        "ranked": rated,
        "unrated": unrated,
    }

    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(
        "// Generated by scripts/import_ratings.py from the ratings spreadsheet.\n"
        "// Re-run that script after updating the sheet; do not edit this by hand.\n"
        "window.RATINGS_DATA = " + json.dumps(payload, indent=2) + ";\n",
        encoding="utf-8",
    )

    print(f"✓ wrote {OUT.relative_to(ROOT)}")
    print(f"  {len(rated)} ranked, {len(unrated)} unrated, "
          f"{payload['ratingEvents']} individual food scores")
    print(f"  best {payload['best']} ({rated[0]['name']}) · "
          f"worst {payload['worst']} ({rated[-1]['name']}) · avg {payload['average']}")
    off = [e['name'] for e in rated if e['tier'] == 'offscale']
    if off:
        print(f"  off-scale (kept as stated): {', '.join(off)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
